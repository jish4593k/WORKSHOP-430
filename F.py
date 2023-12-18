import pandas as pd
from datetime import timedelta, datetime
import random
from openpyxl import load_workbook


def generate_random_timestamps(start_date, end_date):
    date_range = pd.date_range(start=start_date, end=end_date, freq='H')
    random_timestamps = [date + timedelta(hours=random.randint(0, 9)) for date in date_range]
    return random_timestamps


def append_to_excel(file_name, column_name, data):
    try:
        # Load the existing workbook
        book = load_workbook(file_name)
        writer = pd.ExcelWriter(file_name, engine='openpyxl') 
        writer.book = book

      
        pd.DataFrame({column_name: data}).to_excel(writer, index=False, sheet_name='Sheet1', startrow=book['Sheet1'].max_row, header=False)

        
        writer.save()
    except FileNotFoundError:
        
        pd.DataFrame({column_name: data}).to_excel(file_name, index=False)


file_name = "filename.xlsx"

try:
    data = pd.read_excel(file_name)
except FileNotFoundError:
    data = pd.DataFrame()



data['DateAndTimeRecent'] = random_timestamps


append_to_excel(file_name, 'DateAndTimeRecent', random_timestamps)

print(data)
